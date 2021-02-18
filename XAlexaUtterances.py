#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import openpyxl
from pathlib import Path
import re
from parseUtterance import parseUtterance
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

colWidth = 100
fontSize = 14
fontColor = "00008000"

if len(sys.argv) > 1:
    #passing the file as argument
    xlsx_file = sys.argv[1]
elif Path('Utterances.xlsx').is_file():
    #passing the file statically
    print("STATIC MODE OPENING FILE: Utterances.xlsx")
    xlsx_file = Path('Utterances.xlsx')
else:
    print("The Script looks for an 'Utterances.xlsx' file inside the folder where it is launched.")
    print("Alternatively you can provide a suitable .xlsx file as an argument.")
    print("No .xlsx file: Exiting...")
    sys.exit()

wb = openpyxl.load_workbook(xlsx_file) 
# create the output xlsx
wb_out = Workbook()
wb_out.remove(wb_out["Sheet"])

# Read the active sheet or a particular one:
sheet = wb.active
# sheet = wb['it']

# ITERATE ALL SHEETS > ALL COLUMNS > ALL CELLS > PARSE UTTERANCES > GENERATE NEW XLSX WITH ALL PERMUTATIONS
for sheet in wb:
    sheet_out = wb_out.create_sheet(sheet.title)
    sheet_out.column_dimensions["A"].width = colWidth
    sheet_out.column_dimensions["B"].width = colWidth
    print(sheet.title)
    print(sheet.max_row, sheet.max_column)
    
    for col in sheet.iter_cols(1,1):
        for cell in col:
            value = cell.value
            if value is not None:
                # print(value)
                for value in parseUtterance(value):
                    sheet_out.append({1: value, 2:"PASSED"})
    # FONT
    for col in sheet_out.iter_cols(1,2):
         for cell in col:
            cell.font = Font(size=fontSize)
    # COLOR
    for col in sheet_out.iter_cols(2):
         for cell in col:
            cell.font = Font(color=fontColor)
        
print("SAVING XLSX...")
wb_out.save("Utterances_ALL.xlsx")

def test(cell):
    parsed = parseUtterance(cell)
    print(cell)
    print()
    print(*parsed, sep = "\n") 

# test(sheet['A5'].value) 
